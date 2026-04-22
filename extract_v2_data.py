"""
extract_v2_data.py
Baca data stasiun hujan dari PERHITUNGAN_v2.xlsx, hitung ulang semua
kolom hidrologi dengan parameter v2, simpan ke data_grindulu.csv.

Parameter v2 (dari PERHITUNGAN_v2.xlsx):
  CN       = 76.46   (SCS-CN sheet)
  DAS_KM2  = 754.24  (LUAS DAERAH sheet - total Thiessen)
  Baseflow = 1.02111226 m³/s (konstan, dari HUJAN WILAYAH DAS sheet)
  P_DAS    = Thiessen weighted (bukan rata-rata sederhana)

Usage:
    python extract_v2_data.py
"""

import numpy as np
import pandas as pd
import openpyxl
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
XLSX_PATH = Path("PERHITUNGAN_v2.xlsx")
OUT_PATH  = Path("data_grindulu.csv")
SHEET_NAME = "HUJAN WILAYAH DAS (Thiessen Pol"

# ---------------------------------------------------------------------------
# Parameter v2 (dari PERHITUNGAN_v2.xlsx)
# ---------------------------------------------------------------------------
CN         = 76.46
S          = (25400 / CN) - 254     # 78.1998 mm
Ia         = 0.2 * S                # 15.6400 mm
DAS_KM2    = 127.86                 # km² (luas efektif DAS, dari HUJAN WILAYAH DAS sheet)
Q_BASEFLOW = 1.02111226             # m³/s (konstan)

# Bobot Thiessen (dari LUAS DAERAH sheet, total = 754.24 km²)
THIESSEN_WEIGHTS = np.array([
    77.11  / 754.24,   # pacitan
    124.06 / 754.24,   # nawangan
    124.85 / 754.24,   # kebonagung
    117.34 / 754.24,   # bandar
    149.26 / 754.24,   # tegalombo
    161.62 / 754.24,   # tulakan
])
STATION_COLS = ["pacitan", "nawangan", "kebonagung", "bandar", "tegalombo", "tulakan"]


def scs_runoff_mm(p_das: np.ndarray) -> np.ndarray:
    """Hitung hujan efektif Pe (mm) via SCS-CN."""
    pe = np.where(
        p_das > Ia,
        (p_das - Ia) ** 2 / (p_das - Ia + S),
        0.0,
    )
    return pe


def runoff_to_flow(pe_mm: np.ndarray) -> np.ndarray:
    """Konversi kedalaman runoff (mm) ke debit (m³/s)."""
    v_m3 = pe_mm * DAS_KM2 * 1_000   # mm × km² × 1000 = m³
    return v_m3 / 86_400


def main():
    print("=" * 60)
    print("Ekstrak data dari PERHITUNGAN_v2.xlsx")
    print(f"  CN       = {CN}")
    print(f"  S        = {S:.4f} mm")
    print(f"  Ia       = {Ia:.4f} mm")
    print(f"  DAS_KM2  = {DAS_KM2} km²")
    print(f"  Baseflow = {Q_BASEFLOW} m³/s")
    print("=" * 60)

    # --- Baca sheet dari Excel ---
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    dates, stations_data = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            break
        import datetime
        date_val = row[0]
        if isinstance(date_val, datetime.datetime):
            dates.append(date_val.date())
        else:
            dates.append(pd.Timestamp(date_val).date())
        # Kolom 1-6: Pacitan, Nawangan, Kebonagung, Bandar, Tegalombo, Tulakan
        vals = [float(v) if v is not None else 0.0 for v in row[1:7]]
        stations_data.append(vals)

    stations_arr = np.array(stations_data)   # shape (N, 6)

    # --- Hitung kolom hidrologi ---
    p_das    = stations_arr @ THIESSEN_WEIGHTS          # Thiessen weighted (m/s)
    pe       = scs_runoff_mm(p_das)
    q_runoff = runoff_to_flow(pe)
    q_baseflow = np.full(len(dates), Q_BASEFLOW)
    q_total  = q_runoff + q_baseflow

    # --- Susun DataFrame ---
    df = pd.DataFrame({
        "date":       dates,
        **{col: stations_arr[:, i] for i, col in enumerate(STATION_COLS)},
        "p_das":      np.round(p_das, 4),
        "pe":         np.round(pe, 6),
        "q_runoff":   np.round(q_runoff, 6),
        "q_baseflow": q_baseflow,
        "q_total":    np.round(q_total, 6),
    })

    df.to_csv(OUT_PATH, index=False)

    # --- Ringkasan ---
    print(f"\nSaved -> {OUT_PATH}  ({len(df):,} baris)")
    print(f"Periode: {df['date'].iloc[0]} -> {df['date'].iloc[-1]}")
    print(f"\nP_DAS stats (mm/hari):")
    print(f"  mean = {p_das.mean():.3f}  max = {p_das.max():.3f}")
    print(f"\nQ_total stats (m³/s):")
    print(f"  min  = {q_total.min():.6f}")
    print(f"  mean = {q_total.mean():.3f}")
    print(f"  max  = {q_total.max():.3f}")
    print(f"\nVerifikasi P_DAS 2014-01-01: {p_das[0]:.4f} mm  (expected ~ 22.18)")
    print(f"Verifikasi Q_total min       : {q_total.min():.6f}  (expected ~ 1.02111226)")

    Q_MIN_1UNIT = 0.30 * 60.5   # 18.15 m³/s
    operable = (q_total >= Q_MIN_1UNIT).sum()
    print(f"\nHari turbin bisa jalan (Q >= {Q_MIN_1UNIT} m3/s): {operable} / {len(df)} hari ({operable/len(df)*100:.1f}%)")
    print("\nSelesai. Lanjutkan dengan:")
    print("  python train.py --epochs 100")


if __name__ == "__main__":
    main()
